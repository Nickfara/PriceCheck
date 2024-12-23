import json

from openpyxl import load_workbook as open_xlsx

from xlrd import open_workbook as open_xls

# !/usr/bin/env python # -* - coding: utf-8-* -

with open('config.json') as f:
    shops = json.load(f)['shops']

h = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P",
     "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE",
     "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR",
     "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ"]

def read(filename):  # Чтение таблицы
    print(filename)
    format = filename.split('.')[1]
    if True:
        if format == 'xls':  # Чтение xls
            workbook = open_xls('doc/' + str(filename))
            response = xls(workbook)
        elif format == 'xlsx':  # Чтение xlsx
            workbook = open_xlsx('doc/' + str(filename))
            response = xlsx(workbook)
        else:
            response = False  # Формат файла неверный
            error_text = 'Формат файла неверный'
            print(error_text)
    try:
        pass
    except Exception as e:
        response = False
        error_text = 'Сработало исключение:\n' + str(e)
        print(error_text)
    return response


def xls(workbook):
    items = []
    sid = False
    seller = False
    end_time = 0
    for i in shops:
        print(workbook.sheets())
        for i2 in workbook.sheets():
            print(i2)
            worksheet = workbook.sheet_by_index(workbook.sheets().index(i2))
            print('ПОЛНЫЙ потавщик:')
            print(i)
            if i['findtext'].lower() in str(worksheet.cell_value(i['findname'][0], i['findname'][1])).lower():
                sid = i['sid']
                seller = i['seller']
                break

    print(sid)
    if sid:
        for i2 in workbook.sheets():
            worksheet = workbook.sheet_by_index(workbook.sheets().index(i2))
            for i in range(0, worksheet.utter_max_rows):  # Чтение строк
                item = {}
                try:
                    item['seller'] = seller
                    item['name'] = str(worksheet.cell_value(i, sid[0]))
                    if item['name'] != 'None':
                        end_time = 0
                    item['name'] = ''.join(item['name'].split(',')) # Удаление запятых из названия
                    item['cost'] = str(worksheet.cell_value(i, sid[1]))
                    if sid[2] != -1:
                        item['type'] = str(worksheet.cell_value(i, sid[2]))
                    else:
                        item['type'] = ''

                    if item['cost'] == '' or item['cost'] == None:
                        item['cost'] = 'категория'
                    else:
                        items.append(item)
                except:
                    if end_time == 5:
                        print('Конец документа')
                        break
                    end_time += 1
    else:
        items = False
        error_text = 'Не найдено имя поставщика в таблице!'
        print(error_text)

    return items


def xlsx(workbook):
    sheetnames = workbook.sheetnames
    items = []
    sid = False
    seller = False
    end_time = 0
    for i in shops:
        for i2 in sheetnames:
            worksheet = workbook[i2]
            if i['findtext'].lower() in str(worksheet[h[i['findname'][0]]][i['findname'][1]].value).lower():
                sid = (h[i['sid'][0]], h[i['sid'][1]])
                seller = i['seller']
                type_n = (i['sid'][2])
                break
    if sid:
        for i2 in sheetnames:
            print('СТРАНИЦА: ' + str(i2))
            worksheet = workbook[i2]
            for i in range(0, 1000):  # Чтение строк
                item = {}
                try:
                    if 'мороженое' in str(worksheet[f'{sid[0]}'][i].value).lower():
                        print('Дальше мороженое. Остановка чтения файла.')
                        break
                    item['seller'] = seller
                    item['name'] = str(worksheet[f'{sid[0]}'][i].value)
                    if item['name'] != 'None':
                        end_time = 0
                    item['name'] = ''.join(item['name'].split(',')) # Удаление запятых из названия
                    try:
                        item['cost'] = str(worksheet[f'{sid[1]}'][i].value)
                    except:
                        item['cost'] = ''

                    if type_n != -1:
                        try:
                            item['type'] = str(worksheet[f'{sid[2]}'][i].value)
                        except:
                            item['type'] = 'хз'
                    else:
                        item['type'] = ''
                    if item['cost'] == '':
                        item['cost'] = 'категория'
                    items.append(item)
                except:
                    if end_time == 5:
                        print('Конец документа')
                        break
                    end_time += 1
    else:
        items = False
        error_text = 'Не найдено имя поставщика в таблице!'
        print(error_text)
    return items


def filter_names(name):
    name = name.lower()
    base = {
        'куриное': ('цб', 'цыпленка-бройлеров', 'цыпленка бройлера', 'цыпленка'),
        'филе грудки': ('филе  грудки',),
        'грудки куриное': ('кур.грудки',)
    }

    all_check = {}
    find_type = None
    name_ = name

    for i in base:
        for i2 in base[i]:
            all_check[i2] = i  # Наполнение всех вариантов замен


    for i in all_check:
        if i in name:

            find_type = all_check[i]
            break

    if find_type:
        back_list = base[find_type]
        new = find_type

        if type(back_list) not in (list, tuple):
            return name.capitalize()

        for i in back_list:
            if i in name:
                name_ = new.join(name.split(i))
                name_ = name_.capitalize()
                break

    return name_


def scanner(name):
    name = str(name)
    items = []
    result = []
    import os
    for file in shops:
        if file['active']:
            if file['filename'] in os.listdir('doc'):
                table = read(file['filename'])
                if table:
                    for i in table:
                        items.append(i)

    for i in items:
        if name.lower() in i['name'].lower():
            i["name"] = filter_names(i["name"])

            if i['name']:
                result.append(i)

    def key(price):
        try:
            return float(price['cost'])
        except:
            return 0

    result.sort(key=key)
    return result


scanner('')